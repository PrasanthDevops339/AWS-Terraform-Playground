# scripts/get_ec2.py  (FULL FILE — merged with the “rest” you provided, recreated from screenshots)

import json
import boto3
import io
import csv
import os
import time
from botocore.config import Config
from botocore.exceptions import ClientError

import pandas as pd
import logging
import sys
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

from datetime import datetime, timedelta, timezone
from opentelemetry import trace
from opentelemetry.trace import SpanKind

FORMAT = '%(asctime)s - %(levelname)s - [trace_id=%(otelTraceID)s span_id=%(otelSpanID)s resource.service.name=%(otelServiceName)s] - %(message)s'
logger = logging.getLogger(__name__)
logger.setLevel("INFO")
h = logging.StreamHandler(sys.stdout)
h.setFormatter(logging.Formatter(FORMAT))
logger.addHandler(h)

logger.info("Starting EC2")

AGGREGATOR_NAME = os.getenv('AGGREGATOR_NAME')
REGION = os.getenv('REGION')
bucket_name = os.getenv('BUCKET')
dynamo_table = os.getenv('DYNAMO_TABLE')


def main(tries=1):
    # Initialize AWS Config client
    client = boto3.client('config', config=Config(retries={'max_attempts': 10}))

    # Define the query
    query = """
    SELECT
        resourceId,
        accountId,
        resourceType,
        resourceCreationTime,
        configuration.instanceType,
        configuration.imageId,
        availabilityZone,
        configuration.state.name,
        configuration.launchTime,
        configuration.platformDetails,
        configuration.architecture,
        configuration.tags
    WHERE
        resourceType = 'AWS::EC2::Instance'
    ORDER BY
        accountId,
        configuration.platformDetails
    """

    results = []
    try:
        # Initialize the nextToken as None to start the loop
        next_token = None
        # Loop to handle pagination with nextToken
        while True:
            # Execute the query with the current nextToken
            if next_token:
                response = client.select_aggregate_resource_config(
                    Expression=query,
                    ConfigurationAggregatorName=AGGREGATOR_NAME,
                    NextToken=next_token
                )
            else:
                response = client.select_aggregate_resource_config(
                    Expression=query,
                    ConfigurationAggregatorName=AGGREGATOR_NAME
                )

            # Process the results (for example, print them)
            results.extend(response['Results'])

            # Check if there is a nextToken to continue fetching more results
            next_token = response.get('NextToken')

            # If no nextToken, break out of the loop
            if not next_token:
                break

        return results

    except ClientError as e:
        if e.response['Error']['Code'] == 'ThrottlingException':
            if tries <= 3:
                print("Throttling Exception Occured.")
                print("Retrying.....")
                print("Attempt No.: " + str(tries))
                time.sleep(3*tries)
                return main(tries+1)
            else:
                print("Attempted 3 Times But No Success.")
                print("Raising Exception.....")
                raise
        else:
            print(f"Error executing resource query : {str(e)}")
            raise


def get_account_name(account_id):
    try:
        org_client = boto3.client('organizations')
        response = org_client.describe_account(AccountId=account_id)
        account_name = response.get('Account', {}).get('Name')
        return account_name
    except Exception as e:
        logger.error(f"Error getting account name: {e}")
        return None


# A cache to store account names once they've been fetched
account_cache = {}


def get_account_name_cached(account_id):
    # Check if the account name is already cached
    if account_id not in account_cache:
        # If not cached, fetch and store it
        account_cache[account_id] = get_account_name(account_id)
    return account_cache[account_id]


def check_account(account_name):
    # Check if the account is in dynamo to determine 1.0 or 2.0
    dynamodb = boto3.resource('dynamodb', region_name=REGION)
    table = dynamodb.Table(dynamo_table)

    try:
        response = table.query(
            KeyConditionExpression=boto3.dynamodb.conditions.Key('account_name').eq(account_name)
        )

        # Check if the 'Items' list in the response is not empty
        if bool(response.get('Items')):
            return True
        else:
            return False

    except ClientError as e:
        print(f"Error in Dynamo query: {e}")
        return False


def check_ami_age(ami_id, target_region, launch_date):
    try:
        image_name, ami_age, ami_type = "Unknown", "Unknown", "Unknown"
        # Establish an EC2 client connection for the specific region
        ec2_client = boto3.client('ec2', region_name=target_region)

        # Use describe_images to get the AMI details
        response = ec2_client.describe_images(ImageIds=[ami_id])

        if response['Images']:
            creation_date_str = response['Images'][0]['CreationDate']
            image_name = response['Images'][0]['Name']

            ami_info = response['Images'][0]
            ami_type = "None"
            if ami_info.get('Public'):
                ami_type = "Public"
            else:
                ami_type = "Private"

            # Parse the creation date string into a datetime object
            creation_date = datetime.strptime(creation_date_str, "%Y-%m-%dT%H:%M:%S.%fZ").replace(tzinfo=timezone.utc)

            # Calculate the AMI age
            current_time = datetime.now(timezone.utc)
            ami_age = current_time - creation_date
            ami_age = ami_age.days

            return image_name, ami_age, ami_type

        else:
            image_name, ami_age, ami_type = "Unknown", "Unknown", "Unknown"
            print(f"AMI {ami_id} not found in region {target_region}.")
            #calculate from launch time
            current_time = datetime.now(timezone.utc)
            dt_object = datetime.fromisoformat(launch_date.replace('Z', '+00:00'))
            ami_age = current_time - dt_object
            ami_age = ami_age.days
            return '', ami_age, 'Unknown'
            #ami_age = 0
            #return '',ami_age,'Unknown'

    except Exception as e:
        print(f"An error occurred: {e}")


def get_owner(tags):
    for tag in tags:
        if tag['key'] == 'Owner' or 'owner' in tag['key'].lower():
            return tag['value']
    return None


def get_host(tags):
    for tag in tags:
        if tag['key'] == 'Hostname' or 'hostname' in tag['key'].lower():
            return tag['value']
    return None


def get_name(tags):
    for tag in tags:
        if tag['key'] == 'Name':
            return tag['value']
    return None


def check_latest_ami(ami, region):
    ec2 = boto3.client('ec2', region_name=region)
    response = ec2.describe_images(
        Filters=[
            {'Name': 'name', 'Values': [ami]}
        ]
    )

    ami_ver = []
    if response['Images']:
        sorted_images = sorted(response['Images'], key=lambda k: k['CreationDate'], reverse=True)

        # Extract image names and get the top 3
        image_names = [image['Name'] for image in sorted_images[:3]]
        for name in image_names:
            ami_ver.append(name)

    return ami_ver


def get_version_date(ami_name):
    os_version = "Unknown"
    creation_date = "Unknown"

    parts = ami_name.split('-')
    os_version = parts[1]
    if os_version == 'al2':
        os_version = 'Amazon Linux 2'
    elif os_version == 'al2023':
        os_version = 'Amazon Linux 2023'
    elif os_version == 'win22':
        os_version = 'Windows Server 2022'
    elif os_version == 'mla2':
        os_version = 'MarkLogic Amazon Linux 2 ' + parts[2] + '.' + parts[3]
    else:
        os_version = os_version.upper()

    date_string = ami_name.split('T')[0]
    if 'MarkLogic' in os_version:
        date_string = date_string.split('-')[4:]
    else:
        date_string = date_string.split('-')[2:]  # Remove the 'erie' prefix

    date_string = '-'.join(date_string)  # Join the remaining parts with '-'

    # Parse the date string into a datetime object
    ami_date = datetime.strptime(date_string, "%Y-%m-%d")
    creation_date = ami_date.date()

    return os_version, creation_date


def format_sheet(sheet, df, header_format):
    for cell in sheet[1]:
        cell.fill = header_format['fill']
        cell.font = header_format['font']

    for idx, col in enumerate(df.columns, start=1):
        sheet.column_dimensions[get_column_letter(idx)].width = max(
            len(str(col)),
            df[col].astype(str).map(len).max()
        ) + 2


def apply_excel_formatting(writer, ec2_df=None, grouped_df=None):
    workbook = writer.book
    header_format = {
        'fill': PatternFill(start_color='366092', end_color='366092', fill_type='solid'),
        'font': Font(color='FFFFFF', bold=True)
    }

    if ec2_df is not None:
        format_sheet(writer.sheets['EC2 Compliance'], ec2_df, header_format)

    if grouped_df is not None:
        format_sheet(writer.sheets['KPI'], grouped_df, header_format)


def get_patch_details(inst_id):
    ssm_client = boto3.client('ssm')

    try:
        # Example: Filter by a specific resource ID
        response = ssm_client.list_compliance_items(
            ResourceIds=[inst_id],
            ResourceTypes=['ManagedInstance']
        )

        #print("Compliance Items", response)
        for item in response['ComplianceItems']:
            logger.info(f"Resource ID: {item['ResourceId']}")
            logger.info(f"Compliance Type: {item['ComplianceType']}")
            logger.info(f"Compliance Status: {item['Status']}")
            logger.info(f"Severity: {item['Severity']}")
            logger.info(f"Title: {item['Title']}")
            logger.info("-" * 20)
        return response
    except Exception as e:
        logger.error(f"Error retrieving compliance items: {e}")

    try:
        res = ssm_client.describe_instance_patches(
            InstanceId=inst_id
        )
        logger.info("Patches", res)
        for patch in res['Patches']:
            logger.info(
                f"Patch ID: {patch['PatchId']}, State: {patch['State']}, KB ID: {patch.get('KBId', 'N/A')}, | "
                f"Classification: {patch.get('Classification', 'N/A')}"
            )
    except Exception as e:
        logger.error(f"Error retrieving patches for instance {inst_id}: {e}")


if __name__ == '__main__':
    tracer = trace.get_tracer(__name__)
    with tracer.start_as_current_span("GETEC2Values", kind=SpanKind.SERVER):

        results = main()
        filename = datetime.now().strftime("%B")
        output_dir = '/usr/local/bin/aws-scripts/output'
        os.makedirs(output_dir, exist_ok=True)
        excel_path = os.path.join(output_dir, filename)

        data = []

        for item in results:
            parsed_results = json.loads(item)

            resource_id = parsed_results.get('resourceId')
            account_id = parsed_results.get('accountId')
            account_name = get_account_name_cached(account_id)
            resource_type = parsed_results.get('resourceType')
            resource_creationtime = parsed_results.get('resourceCreationTime')
            instance_type = parsed_results.get('configuration', {}).get('instanceType')
            image_id = parsed_results.get('configuration', {}).get('imageId')
            availabilityZone = parsed_results.get('availabilityZone')
            status = parsed_results.get('configuration', {}).get('state', {}).get('name')
            launch_time = parsed_results.get('configuration', {}).get('launchTime')
            platform = parsed_results.get('configuration', {}).get('platformDetails')
            arch = parsed_results.get('configuration', {}).get('architecture')
            tags = parsed_results.get('configuration', {}).get('tags')
            reg = availabilityZone[:-1]

            owner = get_owner(tags)
            host = get_host(tags)
            name = get_name(tags)

            if not check_account(account_name):
                # not 1.0 account

                ami_name, ami_status, ami_type = "Unknown", "Unknown", "Unknown"
                ami_name, ami_status, ami_type = check_ami_age(image_id, reg, launch_time)

                os_version = "Unknown"
                creation_date = "Unknown"

                ami_ver = []
                det = ''

                if int(ami_status) > 90 and ami_type == "Private":
                    ami = ami_name.split("-")[0] + "-" + ami_name.split("-")[1] + "*"
                    ami_ver = check_latest_ami(ami, reg)
                    det = ['Please update your AMI to the latest', ami_ver]

                    # After getting name get version and release date
                    os_version, creation_date = get_version_date(ami_name)

                    patch = get_patch_details(resource_id)
                    logger.info(f"patch details for resource {resource_id} :", patch)

                elif ami_type == 'Public':
                    det = 'Please update your AMI to an Erie Image. Public Images are not supported.'

                elif ami_type == "Unknown":
                    if int(ami_status) > 90:
                        det = 'Please update your AMI to the latest'
                    else:
                        det = ''

                else:
                    # After getting name get version and release date
                    os_version, creation_date = get_version_date(ami_name)

                data.append({
                    'ResourceId': resource_id,
                    'AccountId': account_id,
                    'AccountName': account_name,
                    'ResourceType': resource_type,
                    'CreationTime': resource_creationtime,
                    'InstanceType': instance_type,
                    'ImageId': image_id,
                    'AvailabilityZone': availabilityZone,
                    'Status': status,
                    'LaunchTime': launch_time,
                    'AMIAge': ami_status,
                    'AMIName': ami_name,
                    'AMIType': ami_type,
                    'Platform': platform,
                    'Architecture': arch,
                    'Owner': owner,
                    'HostName': host,
                    'Name': name,
                    'OSVersion': os_version,
                    'CreationDate': creation_date,
                    'Details': det
                })

            else:
                logger.info(f"Account {account_id} & {account_name} are 1.0")

        ec2_df = pd.DataFrame(data)

        grouped_data = []

        grouped_df_public = ec2_df[ec2_df['AMIType'] == 'Public']
        grouped_counts_public = grouped_df_public.groupby(['Platform', 'AMIType']).agg(
            lt_90=('AMIAge', lambda x: (x <= 90).sum()),
            lt_180=('AMIAge', lambda x: ((x > 90) & (x <= 180)).sum()),
            gt_180=('AMIAge', lambda x: (x > 180).sum()),
        ).reset_index()

        grouped_counts_public = grouped_counts_public.rename(columns={
            "lt_90": "AMI's <= 90 days age",
            "lt_180": "AMI's >90 <180days age",
            "gt_180": "AMI's >180 days age"
        })

        for r in grouped_counts_public.to_dict('records'):
            grouped_data.append(r)

        grouped_df_private = ec2_df[ec2_df['AMIType'] == 'Private']
        grouped_counts_private = grouped_df_private.groupby(['Platform', 'AMIType']).agg(
            lt_90=('AMIAge', lambda x: (x <= 90).sum()),
            lt_180=('AMIAge', lambda x: ((x > 90) & (x <= 180)).sum()),
            gt_180=('AMIAge', lambda x: (x > 180).sum()),
        ).reset_index()

        grouped_counts_private = grouped_counts_private.rename(columns={
            "lt_90": "AMI's <= 90 days age",
            "lt_180": "AMI's >90 <180days age",
            "gt_180": "AMI's >180 days age"
        })

        for r in grouped_counts_private.to_dict('records'):
            grouped_data.append(r)

        grouped_df_u = ec2_df[ec2_df['AMIType'] == 'Unknown']
        grouped_counts_u = grouped_df_u.groupby(['Platform', 'AMIType']).agg(
            lt_90=('AMIAge', lambda x: (x <= 90).sum()),
            lt_180=('AMIAge', lambda x: ((x > 90) & (x <= 180)).sum()),
            gt_180=('AMIAge', lambda x: (x > 180).sum()),
        ).reset_index()

        grouped_counts_u = grouped_counts_u.rename(columns={
            "lt_90": "AMI's <= 90 days age",
            "lt_180": "AMI's >90 <180days age",
            "gt_180": "AMI's >180 days age"
        })

        for r in grouped_counts_u.to_dict('records'):
            grouped_data.append(r)

        grouped_df = pd.DataFrame(grouped_data)

        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            if not ec2_df.empty:
                ec2_df.to_excel(writer, sheet_name='EC2 Compliance', index=False)
            if not grouped_df.empty:
                grouped_df.to_excel(writer, sheet_name='KPI', index=False)

            apply_excel_formatting(
                writer,
                ec2_df if not ec2_df.empty else None,
                grouped_df if not grouped_df.empty else None
            )

        # Upload to S3
        current_month_name = datetime.now().strftime("%B")
        object_key = f'ec2_compliance/{current_month_name}_ec2.xlsx'
        s3 = boto3.client('s3', region_name=REGION)
        with open(excel_path, 'rb') as f:
            s3.put_object(
                Bucket=bucket_name,
                Key=object_key,
                Body=f
            )

        logger.info(f"CSV saved to s3://{bucket_name}/{object_key}")
        trace.get_tracer_provider().shutdown()
